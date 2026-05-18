# -*- coding: utf-8 -*-
"""端子表生成器"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# 全局变量：存储从优先级表加载的自定义后缀
_custom_suffixes = []


def parse_plc_points(file_path):
    """解析PLC点位表，同时检查并加载优先级"""
    from openpyxl import load_workbook
    wb = load_workbook(file_path)
    ws = wb.active
    
    rows = list(ws.iter_rows(values_only=True))[1:]  # 跳过表头
    
    i_points = {}  # {设备名_点位名: 地址}
    q_points = {}  # {设备名_点位名: 地址}
    iw_points = {}  # {设备名_点位名: 地址}
    qw_points = {}  # {设备名_点位名: 地址}
    
    for row in rows:
        i_addr, i_name, i_alias, q_addr, q_name, q_alias, iw_addr, iw_name, iw_alias, qw_addr, qw_name, qw_alias = row
        
        if i_name and i_addr:
            i_points[i_name] = i_addr
        if q_name and q_addr:
            q_points[q_name] = q_addr
        if iw_name and iw_addr:
            iw_points[iw_name] = iw_addr
        if qw_name and qw_addr:
            qw_points[qw_name] = qw_addr
    
    # 检查是否有"优先级"Sheet，如果有则自动加载并保存
    if "优先级" in wb.sheetnames:
        priority_dict = load_priority_from_sheet(wb["优先级"])
        if priority_dict:
            save_priority_json(priority_dict)
            update_custom_suffixes(priority_dict)  # 更新自定义后缀
            print(f"从点位表加载了 {len(priority_dict)} 个优先级词条")
    
    return i_points, q_points, iw_points, qw_points


def load_priority_from_sheet(ws):
    """从Excel Sheet加载优先级"""
    priority_dict = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        keyword, order = row[0], row[1]
        if keyword and order is not None:
            try:
                priority_dict[str(keyword).strip()] = int(float(order))
            except:
                pass
    return priority_dict


def save_priority_json(priority_dict, file_path=None):
    """保存优先级到JSON文件"""
    import json
    import os
    
    if file_path is None:
        # 保存在程序目录下
        base_dir = os.path.dirname(os.path.abspath(__file__))
        file_path = os.path.join(base_dir, "priority_config.json")
    
    with open(file_path, 'w', encoding='utf-8') as f:
        json.dump(priority_dict, f, ensure_ascii=False, indent=2)
    
    return file_path


def load_priority_json(file_path=None):
    """从JSON文件加载优先级"""
    import json
    import os
    
    if file_path is None:
        # 查找程序目录
        base_dir = os.path.dirname(os.path.abspath(__file__))
        file_path = os.path.join(base_dir, "priority_config.json")
    
    if os.path.exists(file_path):
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            print(f"加载优先级配置失败: {e}")
    
    return None


def update_custom_suffixes(priority_dict):
    """更新自定义后缀列表（从优先级表关键词中提取）"""
    global _custom_suffixes
    _custom_suffixes = list(priority_dict.keys())
    # 按长度从长到短排序
    _custom_suffixes.sort(key=len, reverse=True)


def extract_device_name(point_name):
    """从点位名提取设备名（去掉点位类型后缀）"""
    if not point_name:
        return None

    name = point_name

    # 常见后缀
    base_suffixes = [
        '备用10', '备用9', '备用8', '备用7', '备用6', '备用5', '备用4', '备用3', '备用2', '备用1',
        '远程', '本地', '运行', '故障', '启动', '停止', '开阀', '关阀', '反馈', '给定',
        '开位', '关位', '公共', '正转', '反转', '使能', '准备好', '报警',
        '频率反馈', '开度反馈', '频率给定', '开度给定',
        'IW', 'QW'
    ]
    
    # 合并自定义后缀（优先级表关键词），放在前面优先匹配
    all_suffixes = _custom_suffixes + base_suffixes
    # 按长度从长到短排序
    all_suffixes.sort(key=len, reverse=True)
    
    for suffix in all_suffixes:
        if name.endswith(suffix):
            base = name[:-len(suffix)]
            # 如果base非空且有意义，返回设备名
            if base and len(base) >= 2:
                return base

    # 默认：设备名是点名去掉最后一个字符
    if len(name) >= 2:
        return name[:-1]
    return name


def group_devices_by_type(i_points, q_points, iw_points, qw_points):
    """按设备类型分组"""
    devices = {}
    
    all_points = []
    for name, addr in i_points.items():
        all_points.append(('I', name, addr))
    for name, addr in q_points.items():
        all_points.append(('Q', name, addr))
    for name, addr in iw_points.items():
        all_points.append(('IW', name, addr))
    for name, addr in qw_points.items():
        all_points.append(('QW', name, addr))
    
    # 按设备名分组
    for ptype, name, addr in all_points:
        device = extract_device_name(name)
        if device:
            if device not in devices:
                devices[device] = {'I': {}, 'Q': {}, 'IW': {}, 'QW': {}}
            devices[device][ptype][name] = addr
    
    return devices



def parse_i_addr(addr):
    """解析I地址为数值用于排序"""
    if not addr:
        return None
    addr = str(addr).strip()
    if not addr.startswith('I'):
        return None
    try:
        parts = addr[1:].split('.')
        return int(parts[0]) * 100 + int(parts[1])  # 字节*100 + 位
    except:
        return None


# 默认优先级表
DEFAULT_PRIORITY = [
    ('远程', 0), ('本地', 1), ('运行', 2), ('开位', 3), ('关位', 4), ('故障', 5),
    ('备用1', 6), ('备用2', 6), ('备用3', 6), ('备用4', 6), ('备用5', 6),
    ('备用6', 6), ('备用7', 6), ('备用8', 6), ('备用9', 6), ('备用10', 6),
    ('公共', 7), ('启动', 8), ('停止', 8), ('开阀', 8), ('关阀', 9), ('反馈', 10), ('给定', 11),
    ('正转', 12), ('反转', 13), ('使能', 14), ('准备好', 15), ('报警', 16)
]


def load_priority_table(file_path=None):
    """加载优先级表
    - 传入了文件路径: 直接从Excel加载（忽略已有JSON），保存到JSON以便后续使用
    - 没传文件路径: 尝试从JSON加载，没有则使用默认
    """
    global _custom_suffixes

    # 1. 如果传入了文件路径，直接从Excel加载
    if file_path is not None:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_path)
            ws = wb.active

            priority_dict = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                keyword, order = row[0], row[1]
                if keyword and order is not None:
                    try:
                        priority_dict[str(keyword).strip()] = int(float(order))
                    except:
                        pass

            if priority_dict:
                save_priority_json(priority_dict)
                update_custom_suffixes(priority_dict)
                print(f"成功加载 {len(priority_dict)} 个优先级词条")
                return priority_dict
        except Exception as e:
            print(f"加载优先级表失败: {e}")

    # 2. 没传文件路径时，从JSON加载（首次加载/后续生成时使用）
    json_priority = load_priority_json()
    if json_priority:
        update_custom_suffixes(json_priority)
        return json_priority

    # 3. 返回默认优先级
    return {name: order for name, order in DEFAULT_PRIORITY}


def export_priority_table(file_path):
    """导出默认优先级表到Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    
    wb = Workbook()
    ws = wb.active
    ws.title = "优先级表"
    
    # 表头
    headers = ['关键词', '优先级(数字越小越靠前)']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # 写入默认数据
    for idx, (keyword, order) in enumerate(DEFAULT_PRIORITY, 2):
        ws.cell(row=idx, column=1, value=keyword)
        ws.cell(row=idx, column=2, value=order)
    
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    
    wb.save(file_path)
    print(f"优先级表已导出: {file_path}")


def _get_sort_key(name, priority_dict=None):
    """获取排序键"""
    if priority_dict is None:
        priority_dict = {name: order for name, order in DEFAULT_PRIORITY}
    for suffix, order in priority_dict.items():
        if name.endswith(suffix):
            return order
    return 999  # 未匹配的放到最后


def _parse_addr(addr):
    """解析地址为数值用于排序"""
    if not addr:
        return (999, 99)
    try:
        if addr.startswith('I') or addr.startswith('Q'):
            prefix = addr[0]
            parts = addr[1:].split('.')
            byte = int(parts[0])
            bit = int(parts[1]) if len(parts) > 1 else 0
            return (byte, bit)
        elif addr.startswith('IW') or addr.startswith('QW'):
            return (int(addr[2:]), 0)
    except:
        pass
    return (999, 99)


def _sorted_keys(pts_dict, priority_dict=None):
    """按类型排序后按地址排序"""
    return sorted(pts_dict.keys(), key=lambda x: (_get_sort_key(x, priority_dict), _parse_addr(pts_dict[x])))


def _sorted_keys_with_polarity(pts_dict, priority_dict=None, plc_type="200SMART"):
    """按类型排序后按地址排序，200SMART模式下先+后-"""
    if plc_type == "200SMART":
        # 200SMART: 按地址顺序，先全部+极性，再全部-极性
        return sorted(pts_dict.keys(), key=lambda x: (_get_sort_key(x, priority_dict), 0 if x.endswith('+') else 1, _parse_addr(pts_dict[x])))
    else:
        return _sorted_keys(pts_dict, priority_dict)


def _parse_ka_ranges(range_str):
    """解析KA电源范围，格式: "1~3,5~7" 返回set"""
    ka_set = set()
    if not range_str or not range_str.strip():
        return ka_set
    for part in range_str.split(','):
        part = part.strip()
        if '~' in part:
            try:
                a, b = part.split('~', 1)
                start, end = int(a.strip()), int(b.strip())
                ka_set.update(range(start, end + 1))
            except:
                pass
    return ka_set


def generate_terminal_rows_with_spacing(devices, add_spacing=True, priority_dict=None,
                                        plc_type="200SMART", power_pos="24V+", power_neg="24V-",
                                        iw_prefix="", qw_prefix="", ka_start=1, ka_power_ranges="",
                                        iw_power_ranges=""):
    """生成端子表行
    add_spacing: 是否在设备之间添加空行
    priority_dict: 优先级字典 {关键词: 优先级}，越小越靠前
    plc_type: "1500" 或 "200SMART"
    power_pos: 24V+标识
    power_neg: 24V-标识
    iw_prefix: IW地址前缀
    qw_prefix: QW地址前缀
    ka_start: KA起始编号
    ka_power_ranges: 需要加24V电源的KA范围，格式 "1~3,5~7"
    """
    rows = []
    ka_power_set = _parse_ka_ranges(ka_power_ranges)
    # 将输入的相对编号转换为绝对KA编号
    if ka_power_set and ka_start != 1:
        ka_power_set = {k + ka_start - 1 for k in ka_power_set}
    iw_power_set = _parse_ka_ranges(iw_power_ranges)
    
    # 使用传入的优先级表，默认使用内置的
    if priority_dict is None:
        priority_dict = {name: order for name, order in DEFAULT_PRIORITY}
    
    def get_min_i_addr(device_points):
        """获取设备I地址的最小值用于排序"""
        i_pts = device_points.get('I', {})
        if not i_pts:
            return None
        min_addr = None
        for addr in i_pts.values():
            val = parse_i_addr(addr)
            if val is not None:
                if min_addr is None or val < min_addr:
                    min_addr = val
        return min_addr
    
    # 按设备类型分组排序
    has_i_devices = []      # 有I的设备
    q_only_devices = []     # 只有Q的设备（无I无IW无QW）
    iw_only_devices = []    # 有IW的设备（无I，可能有Q）
    qw_only_devices = []   # 只有QW的设备（无I无IW，可能有Q）
    
    for device, points in devices.items():
        has_i = bool(points['I'])
        has_q = bool(points['Q'])
        has_iw = bool(points['IW'])
        has_qw = bool(points['QW'])
        
        if has_i:
            # 有I的设备，按I地址排序
            min_addr = get_min_i_addr(points)
            has_i_devices.append((min_addr if min_addr is not None else 999999, device, points))
        elif has_q and not has_iw and not has_qw:
            # 只有Q的设备（无I无IW无QW），放在IW-only之后、QW-only之前
            q_only_devices.append((device, points))
        elif has_iw:
            # 只有IW的设备（可能有Q）
            iw_only_devices.append((device, points))
        elif has_qw:
            # 只有QW的设备（可能有Q）
            qw_only_devices.append((device, points))
    
    # 排序：有I的设备按I地址排序
    has_i_devices.sort(key=lambda x: x[0])
    
    # 按顺序处理各组设备，KA编号全局连续
    ka_counter = [ka_start]  # 使用列表实现可变引用
    
    for item in has_i_devices:
        device = item[1]
        points = item[2]
        _process_device(rows, device, points, ka_counter, add_spacing, plc_type, power_pos, power_neg, priority_dict, iw_prefix, qw_prefix, ka_power_set=ka_power_set, iw_power_set=iw_power_set)

    for device, points in q_only_devices:
        _process_device(rows, device, points, ka_counter, add_spacing, plc_type, power_pos, power_neg, priority_dict, iw_prefix, qw_prefix, ka_power_set=ka_power_set, iw_power_set=iw_power_set)

    for device, points in iw_only_devices:
        _process_device(rows, device, points, ka_counter, add_spacing, plc_type, power_pos, power_neg, priority_dict, iw_prefix, qw_prefix, ka_power_set=ka_power_set, iw_power_set=iw_power_set)

    for device, points in qw_only_devices:
        _process_device(rows, device, points, ka_counter, add_spacing, plc_type, power_pos, power_neg, priority_dict, iw_prefix, qw_prefix, ka_power_set=ka_power_set, iw_power_set=iw_power_set)

    return rows


def _process_device(rows, device, points, ka_counter, add_spacing=True,
                    plc_type="200SMART", power_pos="24V+", power_neg="24V-", priority_dict=None,
                    iw_prefix="", qw_prefix="", ka_power_set=None, iw_power_set=None):
    """处理单个设备，生成端子表行，每个设备序号从1开始
    plc_type: "1500" 或 "200SMART"
    power_pos: 24V+标识
    power_neg: 24V-标识
    priority_dict: 优先级字典
    iw_prefix: IW地址前缀（如 "A" → AIW38+）
    qw_prefix: QW地址前缀（如 "A" → AQW2+）
    """
    i_pts = points['I']
    q_pts = points['Q']
    iw_pts = points['IW']
    qw_pts = points['QW']
    
    has_i = bool(i_pts)
    has_q = bool(q_pts)
    has_iw = bool(iw_pts)
    has_qw = bool(qw_pts)
    
    seq = 1  # 每个设备从序号1开始
    
    # I点位（公共端按优先级插入）
    common_inserted = False
    for key in _sorted_keys(i_pts, priority_dict):
        if not common_inserted and has_i:
            common_key = f"{device}公共"
            if _get_sort_key(common_key, priority_dict) <= _get_sort_key(key, priority_dict):
                rows.append((seq, common_key, power_pos, ""))
                seq += 1
                common_inserted = True
        rows.append((seq, key, i_pts[key], None))
        seq += 1
    
    # 公共端还没插入的话（优先级比所有点都低），放最后
    if has_i and not common_inserted:
        rows.append((seq, f"{device}公共", power_pos, ""))
        seq += 1
    
    # Q点位 - KA继电器（分+和-，+是5脚，-是9脚）
    # KA编号全局连续递增
    if ka_power_set is None:
        ka_power_set = set()
    for key in _sorted_keys(q_pts, priority_dict):
        current_ka = ka_counter[0]
        rows.append((seq, f"{key}+", f"KA{current_ka}_5", None))
        seq += 1
        rows.append((seq, f"{key}-", f"KA{current_ka}_9", None))
        seq += 1
        ka_counter[0] += 1
        # 指定范围内的KA下方加串电正/串电负
        if current_ka in ka_power_set:
            rows.append((seq, f"{device}串电正", power_pos, ""))
            seq += 1
            rows.append((seq, f"{device}串电负", power_neg, ""))
            seq += 1
    
    # IW点位 - 模拟量输入（根据PLC类型不同处理）
    has_iw_row = False
    if plc_type == "1500":
        for iw_idx, key in enumerate(_sorted_keys_with_polarity(iw_pts, priority_dict, plc_type), 1):
            # 1500模式：地址直接显示
            rows.append((seq, key, f"{iw_prefix}{iw_pts[key]}", None))
            seq += 1
            if iw_power_set and iw_idx in iw_power_set:
                rows.append((seq, "二线制串电+", "24V+", ""))
                seq += 1
                rows.append((seq, "二线制串电-", "24V-", ""))
                seq += 1
            has_iw_row = True
    else:
        # 200SMART模式：每个IW点拆分为+和-两行，地址追加极性后缀
        for iw_idx, key in enumerate(_sorted_keys_with_polarity(iw_pts, priority_dict, plc_type), 1):
            addr = iw_pts[key]
            rows.append((seq, f"{key}+", f"{iw_prefix}{addr}+", ""))
            seq += 1
            rows.append((seq, f"{key}-", f"{iw_prefix}{addr}-", ""))
            seq += 1
            if iw_power_set and iw_idx in iw_power_set:
                rows.append((seq, "二线制串电+", "24V+", ""))
                seq += 1
                rows.append((seq, "二线制串电-", "24V-", ""))
                seq += 1
            has_iw_row = True
    if has_iw_row:
        rows.append((seq, "PE端子", "PE", ""))
        seq += 1
    
    # QW点位 - 模拟量输出（分+-）
    has_qw_row = False
    for key in _sorted_keys_with_polarity(qw_pts, priority_dict, plc_type):
        qw_addr = qw_pts[key]
        rows.append((seq, f"{key}+", f"{qw_prefix}{qw_addr}+", None))
        seq += 1
        rows.append((seq, f"{key}-", f"{qw_prefix}{qw_addr}-", None))
        seq += 1
        has_qw_row = True
    if has_qw_row:
        rows.append((seq, "PE端子", "PE", ""))
        seq += 1
    
    if add_spacing:
        rows.append(('', '', '', ''))


def write_sheet(ws, rows):
    """写入数据到工作表"""
    headers = ['序号', '点名', '地址', '备注']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    for idx, row_data in enumerate(rows, 2):
        if len(row_data) == 4:
            seq, name, addr, note = row_data
            ws.cell(row=idx, column=1, value=seq)
            ws.cell(row=idx, column=2, value=name)
            ws.cell(row=idx, column=3, value=addr)
            ws.cell(row=idx, column=4, value=note)
        else:
            ws.cell(row=idx, column=1, value=row_data[0] if len(row_data) > 0 else '')
            ws.cell(row=idx, column=2, value=row_data[1] if len(row_data) > 1 else '')
            ws.cell(row=idx, column=3, value=row_data[2] if len(row_data) > 2 else '')
    
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 30


def generate_terminal_excel(plc_file, output_file, plc_type="200SMART", power_pos="24V+", power_neg="24V-",
                            iw_prefix="", qw_prefix="", ka_start=1, ka_power_ranges="",
                            iw_power_ranges=""):
    """生成端子表 - 生成两个工作表：有空行的和无空行的
    plc_file: PLC点位表Excel文件路径
    output_file: 输出端子表Excel文件路径
    plc_type: "1500" 或 "200SMART"
    power_pos: 24V+标识
    power_neg: 24V-标识
    iw_prefix: IW地址前缀
    qw_prefix: QW地址前缀
    ka_start: KA起始编号（默认1）
    ka_power_ranges: 需要加24V电源的KA范围，格式 "1~3,5~7"
    """
    i_pts, q_pts, iw_pts, qw_pts = parse_plc_points(plc_file)
    
    print(f"I点位: {len(i_pts)}个")
    print(f"Q点位: {len(q_pts)}个")
    print(f"IW点位: {len(iw_pts)}个")
    print(f"QW点位: {len(qw_pts)}个")
    
    devices = group_devices_by_type(i_pts, q_pts, iw_pts, qw_pts)
    print(f"\n设备分组: {len(devices)}个")
    for dev in sorted(devices.keys()):
        print(f"  {dev}: I={len(devices[dev]['I'])} Q={len(devices[dev]['Q'])} IW={len(devices[dev]['IW'])} QW={len(devices[dev]['QW'])}")
    
    # 加载优先级表（从JSON文件或默认）
    priority_dict = load_priority_table()
    has_custom_priority = load_priority_json() is not None
    print(f"\n优先级表: {'已加载自定义' if has_custom_priority else '使用默认'}")
    print(f"PLC类型: {plc_type}, 24V+: {power_pos}, 24V-: {power_neg}")
    
    # 生成两个版本的行数据
    rows_with_spacing = generate_terminal_rows_with_spacing(devices, add_spacing=True, priority_dict=priority_dict,
                                                          plc_type=plc_type, power_pos=power_pos, power_neg=power_neg,
                                                          iw_prefix=iw_prefix, qw_prefix=qw_prefix, ka_start=ka_start,
                                                          ka_power_ranges=ka_power_ranges, iw_power_ranges=iw_power_ranges)
    rows_no_spacing = generate_terminal_rows_with_spacing(devices, add_spacing=False, priority_dict=priority_dict,
                                                          plc_type=plc_type, power_pos=power_pos, power_neg=power_neg,
                                                          iw_prefix=iw_prefix, qw_prefix=qw_prefix, ka_start=ka_start,
                                                          ka_power_ranges=ka_power_ranges, iw_power_ranges=iw_power_ranges)
    
    wb = Workbook()
    
    # 第一个工作表：端子表（有空格）
    ws1 = wb.active
    ws1.title = "端子表"
    write_sheet(ws1, rows_with_spacing)
    print(f"\n[有空行] 端子表: {len(rows_with_spacing)} 行")
    
    # 第二个工作表：端子表紧凑版（无空格）
    ws2 = wb.create_sheet("端子表紧凑版")
    write_sheet(ws2, rows_no_spacing)
    print(f"[无空行] 端子表紧凑版: {len(rows_no_spacing)} 行")
    
    wb.save(output_file)
    print(f"\n端子表已生成: {output_file}")
    print("包含两个工作表: 端子表、端子表紧凑版")


