# -*- coding: utf-8 -*-
"""PLC模块排序生成器"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import re


def generate_i_modules(module_count, start_byte):
    """生成I模块排序
    每组16点: 偶数位在前(0,2,4,6), 奇数位在后(1,3,5,7)
    每个模块占用2个字节
    """
    points = []
    for m in range(module_count):
        module_base = start_byte + m * 2  # 每个模块2字节
        # 字节0: 偶数位0,2,4,6 然后 奇数位1,3,5,7
        for bit in [0, 2, 4, 6]:
            points.append(f"I{module_base}.{bit}")
        for bit in [1, 3, 5, 7]:
            points.append(f"I{module_base}.{bit}")
        # 字节1: 偶数位0,2,4,6 然后 奇数位1,3,5,7
        for bit in [0, 2, 4, 6]:
            points.append(f"I{module_base + 1}.{bit}")
        for bit in [1, 3, 5, 7]:
            points.append(f"I{module_base + 1}.{bit}")
    return points


def generate_q_modules(module_count, start_byte):
    """生成Q模块排序 + KA继电器排列
    每组16点: 偶数位在前, 奇数位在后
    KA每组16个(1-16, 17-32, 33-48...)
    每个模块占用2个字节
    """
    points = []
    for m in range(module_count):
        module_base = start_byte + m * 2  # 每个模块2字节
        # Q点位
        for bit in [0, 2, 4, 6]:
            points.append(f"Q{module_base}.{bit}")
        for bit in [1, 3, 5, 7]:
            points.append(f"Q{module_base}.{bit}")
        for bit in [0, 2, 4, 6]:
            points.append(f"Q{module_base + 1}.{bit}")
        for bit in [1, 3, 5, 7]:
            points.append(f"Q{module_base + 1}.{bit}")
        # KA继电器: 每组16个，第二组17-32，第三组33-48...
        ka_offset = m * 16
        ka_odd = [1, 3, 5, 7, 9, 11, 13, 15]
        ka_even = [2, 4, 6, 8, 10, 12, 14, 16]
        for ka in ka_odd:
            points.append(f"KA{ka + ka_offset}")
        for ka in ka_even:
            points.append(f"KA{ka + ka_offset}")
    return points


def generate_iw_modules(module_count, start_word, plc_type="200SMART", eight_point_count=None):
    """生成IW模块排序
    plc_type: "1500" 或 "200SMART"
    eight_point_count: 前N个模块为8点输入，其余为4点输入。None表示全部为8点。
    8点模块: 每模块16字偏移；4点模块: 每模块8字偏移
    返回: (points, points_per_module) 其中points为每个模块的点数列表
    """
    if eight_point_count is None:
        eight_point_count = module_count

    modules_data = []  # list of lists, each inner list = one module's points
    points_per_module = []

    for m in range(module_count):
        if m < eight_point_count:
            # 8点模块: 16字偏移, 8个通道
            module_base = start_word + m * 16
            if plc_type == "200SMART":
                all_offsets = [0, 2, 4, 6, 8, 10, 12, 14]
                module_pts = []
                for offset in all_offsets:
                    module_pts.append(f"IW{module_base + offset}+")
                    module_pts.append(f"IW{module_base + offset}-")
            else:
                module_pts = []
                for offset in [0, 4, 8, 12]:
                    module_pts.append(f"IW{module_base + offset}")
                for offset in [2, 6, 10, 14]:
                    module_pts.append(f"IW{module_base + offset}")
            points_per_module.append(len(module_pts))
            modules_data.append(module_pts)
        else:
            # 4点模块: 8字偏移, 4个通道
            four_point_base = start_word + eight_point_count * 16
            module_base = four_point_base + (m - eight_point_count) * 8
            if plc_type == "200SMART":
                offsets = [0, 2, 4, 6]
                module_pts = []
                for offset in offsets:
                    module_pts.append(f"IW{module_base + offset}+")
                    module_pts.append(f"IW{module_base + offset}-")
            else:
                module_pts = []
                for offset in [0, 4]:
                    module_pts.append(f"IW{module_base + offset}")
                for offset in [2, 6]:
                    module_pts.append(f"IW{module_base + offset}")
            points_per_module.append(len(module_pts))
            modules_data.append(module_pts)

    # 展平为兼容旧调用的返回格式
    flat_points = []
    for mod in modules_data:
        flat_points.extend(mod)
    return flat_points, points_per_module


def generate_qw_modules(module_count, start_word, plc_type="200SMART"):
    """生成QW模块排序
    每个模块4个字，偏移8
    - 200SMART: 每个通道紧挨着+和- QW10+,QW10-,QW12+,QW12-,...
    - 1500: 偶数组(0,4)先+后-，再奇数组(2,6)先+后-
    """
    points = []
    for m in range(module_count):
        module_base = start_word + m * 8  # 每个模块8字偏移

        if plc_type == "1500":
            # 1500: 偶数组(0,4)先+后-，再奇数组(2,6)先+后-
            for offset in [0, 4]:
                points.append(f"QW{module_base + offset}+")
            for offset in [0, 4]:
                points.append(f"QW{module_base + offset}-")
            for offset in [2, 6]:
                points.append(f"QW{module_base + offset}+")
            for offset in [2, 6]:
                points.append(f"QW{module_base + offset}-")
        else:
            # 200SMART: 每个通道紧挨着+和-
            for offset in [0, 2, 4, 6]:
                points.append(f"QW{module_base + offset}+")
                points.append(f"QW{module_base + offset}-")
    return points


def generate_module_excel(i_modules, i_start, q_modules, q_start,
                          iw_modules, iw_start, qw_modules, qw_start, output_file,
                          plc_type="200SMART", iw_eight_count=None):
    """生成模块排序Excel - 每个模块一列
    plc_type: "1500" 或 "200SMART"
    iw_eight_count: 前N个IW模块为8点，剩余为4点。None表示全部8点。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "PLC模块排序"

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    center_align = Alignment(horizontal='center', vertical='center')

    # 标题行
    ws.merge_cells('A1:F1')
    ws['A1'] = f"PLC模块排序表 ({plc_type})"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center_align

    # 生成各模块数据
    i_data = generate_i_modules(i_modules, i_start)  # 每个模块16点
    q_data = generate_q_modules(q_modules, q_start)  # 每个模块32点(Q16+KA16)
    iw_flat, iw_pts_per_module = generate_iw_modules(iw_modules, iw_start, plc_type, iw_eight_count)
    qw_data = generate_qw_modules(qw_modules, qw_start, plc_type)  # QW点数根据plc_type变化

    # 确定每列的点数
    i_points_per_module = 16
    q_points_per_module = 32
    qw_points_per_module = 8  # QW每模块8点(4正+4负)，1500与SMART规则相同

    # I模块 - 每模块1列
    col = 1
    ws.cell(row=3, column=col, value="").font = header_font
    ws.cell(row=3, column=col).fill = header_fill

    for m in range(i_modules):
        col = m + 2
        title_cell = ws.cell(row=3, column=col, value=f"I模块{m+1}")
        title_cell.font = header_font
        title_cell.fill = header_fill
        title_cell.alignment = center_align
        for i in range(i_points_per_module):
            idx = m * i_points_per_module + i
            ws.cell(row=i+4, column=col, value=i_data[idx])

    # Q模块 - 每模块1列
    q_start_col = 2 + i_modules
    for m in range(q_modules):
        col = q_start_col + m
        title_cell = ws.cell(row=3, column=col, value=f"Q模块{m+1}")
        title_cell.font = header_font
        title_cell.fill = header_fill
        title_cell.alignment = center_align
        for i in range(q_points_per_module):
            idx = m * q_points_per_module + i
            ws.cell(row=i+4, column=col, value=q_data[idx])

    # IW模块 - 每模块1列，每列点数可能不同
    iw_start_col = q_start_col + q_modules
    for m in range(iw_modules):
        col = iw_start_col + m
        pts_count = iw_pts_per_module[m]
        title_cell = ws.cell(row=3, column=col, value=f"IW模块{m+1}")
        title_cell.font = header_font
        title_cell.fill = header_fill
        title_cell.alignment = center_align
        # 计算该模块在展平列表中的起始索引
        flat_idx = sum(iw_pts_per_module[:m])
        for i in range(pts_count):
            ws.cell(row=i+4, column=col, value=iw_flat[flat_idx + i])

    # QW模块 - 每模块1列
    qw_start_col = iw_start_col + iw_modules
    for m in range(qw_modules):
        col = qw_start_col + m
        title_cell = ws.cell(row=3, column=col, value=f"QW模块{m+1}")
        title_cell.font = header_font
        title_cell.fill = header_fill
        title_cell.alignment = center_align
        for i in range(qw_points_per_module):
            idx = m * qw_points_per_module + i
            ws.cell(row=i+4, column=col, value=qw_data[idx])

    # 设置列宽
    for c in range(1, qw_start_col + qw_modules):
        ws.column_dimensions[chr(64+c) if c <= 26 else 'A' + chr(64+c-26)].width = 12

    # 添加说明
    max_iw = max(iw_pts_per_module) if iw_pts_per_module else 0
    note_row = 4 + max(i_points_per_module, q_points_per_module, max_iw, qw_points_per_module) + 1
    ws.cell(row=note_row, column=1, value="说明:").font = header_font
    ws.cell(row=note_row, column=2, value=f"I模块: 每模块{i_points_per_module}点 ({i_modules}个)")
    ws.cell(row=note_row+1, column=2, value=f"Q模块: 每模块{q_points_per_module}点(Q16+KA16) ({q_modules}个)")
    iw_8 = min(iw_modules, iw_eight_count if iw_eight_count is not None else iw_modules)
    iw_4 = max(0, iw_modules - iw_8)
    ws.cell(row=note_row+2, column=2, value=f"IW模块: {iw_8}个8点 + {iw_4}个4点 = {sum(iw_pts_per_module)}点")
    ws.cell(row=note_row+3, column=2, value=f"QW模块: 每模块{qw_points_per_module}字 ({qw_modules}个)")

    wb.save(output_file)
    print(f"模块排序表已生成: {output_file}")
    print(f"I: {len(i_data)}点 | Q: {len(q_data)}点 | IW: {sum(iw_pts_per_module)}点 | QW: {len(qw_data)}点")


