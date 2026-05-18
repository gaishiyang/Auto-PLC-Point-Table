# -*- coding: utf-8 -*-
"""EPLAN模块点位分配表生成工具"""

def generate_i_modules(module_count, start_byte):
    """生成I模块点位分配
    每组16点，偶数位在前，奇数位在后
    """
    addresses = []
    current_byte = start_byte

    for module in range(module_count):
        module_addr = []
        # 偶数位: .0, .2, .4, .6 (字节内), 然后下一字节
        for bit in [0, 2, 4, 6]:
            module_addr.append(f"I{current_byte}.{bit}")
        current_byte += 1
        for bit in [0, 2, 4, 6]:
            module_addr.append(f"I{current_byte}.{bit}")
        current_byte += 1
        # 奇数位: .1, .3, .5, .7
        for bit in [1, 3, 5, 7]:
            module_addr.append(f"I{current_byte}.{bit}")
        current_byte += 1
        for bit in [1, 3, 5, 7]:
            module_addr.append(f"I{current_byte}.{bit}")
        current_byte += 1

        addresses.append(module_addr)

    return addresses


def generate_q_modules(module_count, start_byte):
    """生成Q模块点位分配 + KA继电器
    每组16点，偶数位在前，奇数位在后
    """
    addresses = []
    current_byte = start_byte

    for module in range(module_count):
        module_addr = []
        # 偶数位
        for bit in [0, 2, 4, 6]:
            module_addr.append(f"Q{current_byte}.{bit}")
        current_byte += 1
        for bit in [0, 2, 4, 6]:
            module_addr.append(f"Q{current_byte}.{bit}")
        current_byte += 1
        # 奇数位
        for bit in [1, 3, 5, 7]:
            module_addr.append(f"Q{current_byte}.{bit}")
        current_byte += 1
        for bit in [1, 3, 5, 7]:
            module_addr.append(f"Q{current_byte}.{bit}")
        current_byte += 1

        addresses.append(module_addr)

    # 添加KA继电器列表
    ka_list = [f"KA{i}" for i in range(1, module_count * 16 + 1)]

    return addresses, ka_list


def generate_iw_modules(module_count, start_word, eight_point_count=None):
    """生成IW模块点位分配
    支持8点/4点混合: 前eight_point_count个为8点模块，其余为4点模块
    8点: 每组8点，交错间隔4
    4点: 每组4点，交错间隔4
    """
    if eight_point_count is None:
        eight_point_count = module_count

    addresses = []

    for m in range(module_count):
        if m < eight_point_count:
            # 8点模块: 16字偏移
            current = start_word + m * 16
            module_addr = []
            for offset in [0, 4, 8, 12]:
                module_addr.append(f"IW{current + offset}")
            current += 2
            for offset in [2, 6, 10, 14]:
                module_addr.append(f"IW{current + offset - 2}")
            current += 2
        else:
            # 4点模块: 8字偏移
            four_point_base = start_word + eight_point_count * 16
            current = four_point_base + (m - eight_point_count) * 8
            module_addr = []
            for offset in [0, 4]:
                module_addr.append(f"IW{current + offset}")
            current += 2
            for offset in [2, 6]:
                module_addr.append(f"IW{current + offset - 2}")
            current += 2

        addresses.append(module_addr)

    return addresses


def generate_qw_modules(module_count, start_word):
    """生成QW模块点位分配
    每组4点，交错
    顺序: 起始, +4, +2, +6
    """
    addresses = []
    current = start_word

    for module in range(module_count):
        module_addr = [f"QW{current}"]
        current += 4
        module_addr.append(f"QW{current}")
        current -= 2
        module_addr.append(f"QW{current}")
        current += 4
        module_addr.append(f"QW{current}")
        current += 2

        addresses.append(module_addr)

    return addresses


def generate_eplan_sheet(module_i, start_i, module_q, start_q, module_iw, start_iw, module_qw, start_qw, output_file, iw_eight_count=None):
    """生成EPLAN模块点位表"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "EPLAN点位"

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 生成数据
    i_data = generate_i_modules(module_i, start_i)
    q_data, ka_list = generate_q_modules(module_q, start_q)
    iw_data = generate_iw_modules(module_iw, start_iw, iw_eight_count)
    qw_data = generate_qw_modules(module_qw, start_qw)

    # 计算最大行数
    max_points = max(
        len(i_data[0]) if i_data else 0,
        len(q_data[0]) if q_data else 0,
        max((len(d) for d in iw_data), default=0),
        len(qw_data[0]) if qw_data else 0,
        len(ka_list)
    )

    # 第1行标题
    ws.cell(row=1, column=1, value="EPLAN模块点位分配表")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

    # I模块
    start_row = 3
    ws.cell(row=start_row, column=1, value=f"I模块 (共{module_i}个, 起始I{start_i}, 共{module_i*16}点)")
    ws.cell(row=start_row, column=1).font = Font(bold=True)

    for col, module_data in enumerate(i_data, 2):
        ws.cell(row=start_row, column=col, value=f"模块{col-1}")
        ws.cell(row=start_row, column=col).font = Font(bold=True)
        for row, addr in enumerate(module_data, start_row + 1):
            ws.cell(row=row, column=col, value=addr)

    # Q模块
    start_row = max_points + start_row + 3
    ws.cell(row=start_row, column=1, value=f"Q模块 (共{module_q}个, 起始Q{start_q}, 共{module_q*16}点)")
    ws.cell(row=start_row, column=1).font = Font(bold=True)

    for col, module_data in enumerate(q_data, 2):
        ws.cell(row=start_row, column=col, value=f"模块{col-1}")
        ws.cell(row=start_row, column=col).font = Font(bold=True)
        for row, addr in enumerate(module_data, start_row + 1):
            ws.cell(row=row, column=col, value=addr)

    # KA继电器
    start_row_ka = start_row
    ws.cell(row=start_row_ka, column=module_q + 3, value="KA继电器")
    ws.cell(row=start_row_ka, column=module_q + 3).font = Font(bold=True)

    for row, ka in enumerate(ka_list, start_row_ka + 1):
        ws.cell(row=row, column=module_q + 3, value=ka)

    # IW模块
    start_row = start_row + max_points + 3
    iw_total = sum(len(d) for d in iw_data)
    ws.cell(row=start_row, column=1, value=f"IW模块 (共{module_iw}个, 起始IW{start_iw}, 共{iw_total}点)")
    ws.cell(row=start_row, column=1).font = Font(bold=True)

    for col, module_data in enumerate(iw_data, 2):
        ws.cell(row=start_row, column=col, value=f"模块{col-1}")
        ws.cell(row=start_row, column=col).font = Font(bold=True)
        for row, addr in enumerate(module_data, start_row + 1):
            ws.cell(row=row, column=col, value=addr)

    # QW模块
    start_row = start_row + max(len(d) for d in iw_data) + 3
    ws.cell(row=start_row, column=1, value=f"QW模块 (共{module_qw}个, 起始QW{start_qw}, 共{module_qw*4}点)")
    ws.cell(row=start_row, column=1).font = Font(bold=True)

    for col, module_data in enumerate(qw_data, 2):
        ws.cell(row=start_row, column=col, value=f"模块{col-1}")
        ws.cell(row=start_row, column=col).font = Font(bold=True)
        for row, addr in enumerate(module_data, start_row + 1):
            ws.cell(row=row, column=col, value=addr)

    # 调整列宽
    for col in range(1, module_q + 5):
        ws.column_dimensions[chr(64 + col) if col <= 26 else 'A' + chr(64 + col - 26)].width = 12

    wb.save(output_file)
    print(f"EPLAN点位表已生成: {output_file}")
    print(f"  I模块: {module_i}个 x 16点 = {module_i*16}点")
    print(f"  Q模块: {module_q}个 x 16点 = {module_q*16}点")
    print(f"  IW模块: {module_iw}个 = {iw_total}点")
    print(f"  QW模块: {module_qw}个 x 4点 = {module_qw*4}点")
