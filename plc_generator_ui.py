# -*- coding: utf-8 -*-
"""PLC点位表自动生成工具 - 整合版"""
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from generate_plc import parse_template, expand_points, octal_bit, octal_word, generate_var_name, generate_alias, export_alias_map, load_alias_map
from generate_module import generate_module_excel, generate_i_modules, generate_q_modules, generate_iw_modules, generate_qw_modules
from generate_terminal import generate_terminal_excel, parse_plc_points, group_devices_by_type, generate_terminal_rows_with_spacing, load_priority_table, load_priority_json, DEFAULT_PRIORITY
from generate_eplan import generate_eplan_sheet


class PLCGeneratorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("PLC点位表自动生成工具")
        self.root.geometry("1100x760")
        self.root.minsize(960, 640)

        # 配色方案 (Tailwind-inspired)
        self.c = {
            'sd_bg': '#0f172a',      # sidebar 背景
            'sd_hov': '#1e293b',     # sidebar hover
            'sd_act': '#3b82f6',     # sidebar 激活指示
            'sd_txt': '#64748b',     # sidebar 文字
            'sd_txt_a': '#f8fafc',   # sidebar 激活文字
            'bg': '#f1f5f9',         # 主背景
            'card': '#ffffff',       # 卡片背景
            'bd': '#e2e8f0',         # 边框
            'accent': '#3b82f6',     # 强调色
            'suc': '#10b981',        # 成功
            'p1': '#1e293b',         # 主要文字
            'p2': '#64748b',         # 次要文字
            'hdr_bg': '#ffffff',     # 头部背景
        }

        self.root.configure(bg=self.c['bg'])

        # 当前页面
        self.current_page = None
        self.pages = {}
        self.nav_btns = {}
        self.status_var = tk.StringVar(value="就绪")

        self.setup_styles()
        self.build_ui()
        self.show_page('point')

    # ── 样式 ──────────────────────────────────────────────
    def setup_styles(self):
        style = ttk.Style()
        style.theme_use('clam')
        style.configure('.', font=('Microsoft YaHei', 10))
        style.configure('Card.TLabelframe', background=self.c['card'])
        style.configure('Card.TLabelframe.Label', background=self.c['card'],
                        foreground=self.c['p1'], font=('Microsoft YaHei', 10, 'bold'))
        style.configure('TEntry', fieldbackground=self.c['card'])
        style.configure('TSpinbox', fieldbackground=self.c['card'])
        style.map('TEntry', fieldbackground=[('focus', '#ffffff')])
        style.configure('TFrame', background=self.c['bg'])
        style.configure('TLabel', background=self.c['bg'], foreground=self.c['p1'])
        style.configure('TLabelframe', background=self.c['bg'])
        style.configure('TLabelframe.Label', background=self.c['bg'], foreground=self.c['p1'])
        style.configure('TRadiobutton', background=self.c['card'], foreground=self.c['p1'])
        style.configure('Treeview', background=self.c['card'], fieldbackground=self.c['card'],
                        foreground=self.c['p1'], rowheight=28)
        style.configure('Treeview.Heading', font=('Microsoft YaHei', 9, 'bold'),
                        background='#f8fafc', foreground=self.c['p1'], relief='flat')
        style.map('Treeview.Heading', background=[('active', '#f1f5f9')])

    # ── 主布局 ────────────────────────────────────────────
    def build_ui(self):
        root = tk.Frame(self.root, bg=self.c['bg'])
        root.pack(fill=tk.BOTH, expand=True)

        # 顶部标题栏
        hdr = tk.Frame(root, bg=self.c['hdr_bg'], height=52)
        hdr.pack(fill=tk.X)
        hdr.pack_propagate(False)
        tk.Frame(hdr, bg=self.c['bd'], height=1).pack(side=tk.BOTTOM, fill=tk.X)

        tk.Label(hdr, text="PLC点位表自动生成工具", font=('Microsoft YaHei', 15, 'bold'),
                 bg=self.c['hdr_bg'], fg=self.c['p1']).pack(side=tk.LEFT, padx=24, pady=10)

        # 主体
        body = tk.Frame(root, bg=self.c['bg'])
        body.pack(fill=tk.BOTH, expand=True)

        self.build_sidebar(body)

        sep = tk.Frame(body, bg=self.c['bd'], width=1)
        sep.pack(side=tk.LEFT, fill=tk.Y)

        # 页面容器
        self.page_container = tk.Frame(body, bg=self.c['bg'])
        self.page_container.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # 状态栏
        st = tk.Frame(root, bg=self.c['hdr_bg'], height=26)
        st.pack(fill=tk.X)
        st.pack_propagate(False)
        tk.Frame(st, bg=self.c['bd'], height=1).pack(side=tk.TOP, fill=tk.X)
        tk.Label(st, textvariable=self.status_var, font=('Microsoft YaHei', 9),
                 bg=self.c['hdr_bg'], fg=self.c['p2']).pack(side=tk.LEFT, padx=14, pady=2)

    # ── 侧边栏 ────────────────────────────────────────────
    def build_sidebar(self, parent):
        sd = tk.Frame(parent, bg=self.c['sd_bg'], width=210)
        sd.pack(side=tk.LEFT, fill=tk.Y)
        sd.pack_propagate(False)

        tk.Label(sd, text="功能导航", font=('Microsoft YaHei', 11, 'bold'),
                 bg=self.c['sd_bg'], fg='#94a3b8').pack(padx=20, pady=(22, 14), anchor=tk.W)

        nav = [
            ('point',   '点位表生成'),
            ('module',  '模块排序生成'),
            ('terminal','端子表生成'),
            ('eplan',   'EPLAN点位表'),
        ]

        for pid, label in nav:
            # 按钮容器
            bf = tk.Frame(sd, bg=self.c['sd_bg'], cursor='hand2')
            bf.pack(fill=tk.X, padx=8, pady=2)

            # 激活指示条
            ind = tk.Frame(bf, bg=self.c['sd_bg'], width=3)
            ind.pack(side=tk.LEFT, fill=tk.Y)

            # 图标圆点 + 文字
            dot = tk.Label(bf, text='●', font=('Microsoft YaHei', 8),
                           bg=self.c['sd_bg'], fg=self.c['sd_txt'])
            dot.pack(side=tk.LEFT, padx=(8, 6))

            lbl = tk.Label(bf, text=label, font=('Microsoft YaHei', 10),
                           bg=self.c['sd_bg'], fg=self.c['sd_txt'],
                           anchor=tk.W, padx=0, pady=10)
            lbl.pack(side=tk.LEFT, fill=tk.X, expand=True)

            # 绑定交互
            for w in (bf, dot, lbl):
                w.bind('<Button-1>', lambda e, p=pid: self.show_page(p))
            bf.bind('<Enter>', lambda e, f=bf: f.configure(bg=self.c['sd_hov']) or all(
                c.configure(bg=self.c['sd_hov']) for c in f.winfo_children()))
            bf.bind('<Leave>', lambda e, f=bf: (
                setattr(f, '_bg', self.c['sd_bg']),
                f.configure(bg=self.c['sd_bg']),
                [c.configure(bg=self.c['sd_bg']) for c in f.winfo_children() if c != ind]
            ) if self.nav_btns.get(pid, [None])[0] != f else None)

            self.nav_btns[pid] = (bf, ind)

    # ── 页面切换 ──────────────────────────────────────────
    def show_page(self, pid):
        if self.current_page == pid:
            return
        self.current_page = pid

        # 更新导航状态
        for k, (bf, ind) in self.nav_btns.items():
            active = k == pid
            bg = self.c['sd_act'] if active else self.c['sd_bg']
            fg = self.c['sd_txt_a'] if active else self.c['sd_txt']
            ind.configure(bg=bg)
            for c in bf.winfo_children():
                if isinstance(c, tk.Label):
                    c.configure(fg=fg)

        # 隐藏当前页面
        for w in self.page_container.winfo_children():
            w.pack_forget()

        # 如果页面已缓存，直接显示
        if pid in self.pages:
            self.pages[pid].pack(fill=tk.BOTH, expand=True)
            return

        # 构建新页面
        build_map = {
            'point': self._build_point_page,
            'module': self._build_module_page,
            'terminal': self._build_terminal_page,
            'eplan': self._build_eplan_page,
        }
        build_map[pid]()

    # ══════════════════════════════════════════════════════
    # 页面1: 点位表生成
    # ══════════════════════════════════════════════════════
    def _build_point_page(self):
        self.status_var.set("就绪")
        f = ttk.Frame(self.page_container, padding=20)
        f.pack(fill=tk.BOTH, expand=True)
        self.pages['point'] = f

        # ── 文件设置 ──
        gf1 = ttk.LabelFrame(f, text="文件设置", padding=12, style='Card.TLabelframe')
        gf1.pack(fill=tk.X, pady=(0, 10))

        r0 = ttk.Frame(gf1)
        r0.pack(fill=tk.X, pady=3)
        ttk.Label(r0, text="模板文件:").pack(side=tk.LEFT)
        self.template_path = tk.StringVar(value="点表模板.xlsx")
        ttk.Entry(r0, textvariable=self.template_path, width=45).pack(side=tk.LEFT, padx=8)
        ttk.Button(r0, text="浏览", command=self.browse_template).pack(side=tk.LEFT)

        r1 = ttk.Frame(gf1)
        r1.pack(fill=tk.X, pady=3)
        ttk.Label(r1, text="输出文件:").pack(side=tk.LEFT)
        self.output_path = tk.StringVar(value="PLC点位表.xlsx")
        ttk.Entry(r1, textvariable=self.output_path, width=45).pack(side=tk.LEFT, padx=8)
        ttk.Button(r1, text="浏览", command=self.browse_output).pack(side=tk.LEFT)

        # ── 别名与优先级 ──
        gf2 = ttk.LabelFrame(f, text="别名映射表 (可选)", padding=12, style='Card.TLabelframe')
        gf2.pack(fill=tk.X, pady=(0, 10))

        r2 = ttk.Frame(gf2)
        r2.pack(fill=tk.X, pady=3)
        ttk.Label(r2, text="别名映射文件:").pack(side=tk.LEFT)
        self.alias_file = tk.StringVar()
        ttk.Entry(r2, textvariable=self.alias_file, width=40).pack(side=tk.LEFT, padx=8)
        ttk.Button(r2, text="导入", command=self.import_alias).pack(side=tk.LEFT, padx=2)
        ttk.Button(r2, text="导出默认", command=self.export_alias).pack(side=tk.LEFT, padx=2)
        ttk.Label(r2, text="自定义设备类型对应的别名模板", font=('Microsoft YaHei', 9),
                  foreground=self.c['p2']).pack(side=tk.LEFT, padx=12)

        gf3 = ttk.LabelFrame(f, text="优先级表 (可选, 用于排序规则)", padding=12, style='Card.TLabelframe')
        gf3.pack(fill=tk.X, pady=(0, 10))

        r3 = ttk.Frame(gf3)
        r3.pack(fill=tk.X, pady=3)
        ttk.Label(r3, text="优先级文件:").pack(side=tk.LEFT)
        self.point_priority_file = tk.StringVar()
        ttk.Entry(r3, textvariable=self.point_priority_file, width=40).pack(side=tk.LEFT, padx=8)
        ttk.Button(r3, text="导入", command=self.import_priority_point).pack(side=tk.LEFT, padx=2)
        ttk.Button(r3, text="导出默认", command=self.export_priority_point).pack(side=tk.LEFT, padx=2)

        # ── 起始地址 + 操作按钮 ──
        row_btn = ttk.Frame(f)
        row_btn.pack(fill=tk.X, pady=(0, 10))

        ttk.Button(row_btn, text="刷新预览", command=self.refresh_preview).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(row_btn, text="生成点位表", command=self.generate_plc, width=15).pack(side=tk.LEFT)

        gf4 = ttk.LabelFrame(f, text="起始地址设置", padding=12, style='Card.TLabelframe')
        gf4.pack(fill=tk.X, pady=(0, 10))


        self.i_start = tk.IntVar(value=2)
        self.q_start = tk.IntVar(value=0)
        self.iw_start = tk.IntVar(value=10)
        self.qw_start = tk.IntVar(value=0)

        addr_grid = ttk.Frame(gf4)
        addr_grid.pack(fill=tk.X)
        for i, (txt, var, note) in enumerate([
            ("I地址起始字节:", self.i_start, "I字节.位"),
            ("Q地址起始字节:", self.q_start, "Q字节.位"),
            ("IW地址起始偏移:", self.iw_start, "IW偏移"),
            ("QW地址起始偏移:", self.qw_start, "QW偏移"),
        ]):
            ttk.Label(addr_grid, text=txt).grid(row=i//2, column=i%2*3, sticky=tk.W, pady=2, padx=(0, 4))
            ttk.Spinbox(addr_grid, from_=0, to=1000, textvariable=var, width=8).grid(row=i//2, column=i%2*3+1, sticky=tk.W, padx=(0, 4))
            ttk.Label(addr_grid, text=note, foreground=self.c['p2']).grid(row=i//2, column=i%2*3+2, sticky=tk.W, padx=(0, 24))

        # ── 设备类型预览 ──
        gf5 = ttk.LabelFrame(f, text="设备类型预览", padding=8, style='Card.TLabelframe')
        gf5.pack(fill=tk.BOTH, expand=True)

        cols = ("设备类型", "设备数", "设备名称", "I", "Q", "IW", "QW")
        self.preview_tree = ttk.Treeview(gf5, columns=cols, show="headings", height=7)
        widths = [80, 60, 150, 80, 80, 60, 60]
        for col, w in zip(cols, widths):
            self.preview_tree.heading(col, text=col)
            self.preview_tree.column(col, width=w, anchor=tk.CENTER)
        scroll = ttk.Scrollbar(gf5, orient=tk.VERTICAL, command=self.preview_tree.yview)
        self.preview_tree.configure(yscrollcommand=scroll.set)
        self.preview_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scroll.pack(side=tk.RIGHT, fill=tk.Y)

        self.refresh_preview()

    # ── 点位表页面方法 ──
    def browse_template(self):
        fn = filedialog.askopenfilename(title="选择模板", filetypes=[("Excel", "*.xlsx")])
        if fn:
            self.template_path.set(fn)
            self.refresh_preview()

    def browse_output(self):
        fn = filedialog.asksaveasfilename(title="保存位置", defaultextension=".xlsx",
                                          filetypes=[("Excel", "*.xlsx")])
        if fn:
            self.output_path.set(fn)

    def import_alias(self):
        fn = filedialog.askopenfilename(title="导入别名映射表", filetypes=[("Excel", "*.xlsx")])
        if fn:
            try:
                load_alias_map(fn)
                self.alias_file.set(fn)
                messagebox.showinfo("成功", f"已导入别名映射表:\n{fn}")
            except Exception as e:
                messagebox.showerror("错误", f"导入失败:\n{str(e)}")

    def export_alias(self):
        fn = filedialog.asksaveasfilename(title="导出别名映射表", defaultextension=".xlsx",
                                          initialfile="别名映射表.xlsx", filetypes=[("Excel", "*.xlsx")])
        if fn:
            try:
                export_alias_map(fn)
                self.alias_file.set(fn)
                messagebox.showinfo("成功", f"已导出别名映射表:\n{fn}\n\n可编辑后导入使用")
            except Exception as e:
                messagebox.showerror("错误", str(e))

    def import_priority_point(self):
        fn = filedialog.askopenfilename(title="导入优先级表", filetypes=[("Excel", "*.xlsx")])
        if fn:
            try:
                from generate_terminal import save_priority_json
                priority_dict = load_priority_table(fn)
                if priority_dict:
                    save_priority_json(priority_dict)
                self.point_priority_file.set(fn)
                messagebox.showinfo("成功", f"已导入优先级表:\n{fn}")
            except Exception as e:
                messagebox.showerror("错误", f"导入失败:\n{str(e)}")

    def export_priority_point(self):
        from generate_terminal import export_priority_table
        fn = filedialog.asksaveasfilename(title="导出优先级表", defaultextension=".xlsx",
                                          initialfile="优先级表.xlsx", filetypes=[("Excel", "*.xlsx")])
        if fn:
            try:
                export_priority_table(fn)
                self.point_priority_file.set(fn)
                messagebox.showinfo("成功", f"已导出优先级表:\n{fn}\n\n可编辑后导入使用")
            except Exception as e:
                messagebox.showerror("错误", str(e))

    def refresh_preview(self):
        for item in self.preview_tree.get_children():
            self.preview_tree.delete(item)
        path = self.template_path.get()
        if not os.path.exists(path):
            self.status_var.set(f"文件不存在: {path}")
            return
        try:
            device_definitions = parse_template(path)
            for dtype, defs in device_definitions.items():
                count = len(defs['devices'])
                devices_str = ",".join(defs['devices'][:4])
                if len(defs['devices']) > 4:
                    devices_str += "..."
                self.preview_tree.insert("", tk.END, values=(
                    dtype, count, devices_str,
                    ",".join(defs['i_templates']) if defs['i_templates'] else "-",
                    ",".join(defs['q_templates']) if defs['q_templates'] else "-",
                    ",".join(defs['iw_templates']) if defs['iw_templates'] else "-",
                    ",".join(defs['qw_templates']) if defs['qw_templates'] else "-"
                ))
            self.status_var.set(f"就绪 - 共 {len(device_definitions)} 种设备类型")
        except Exception as e:
            self.status_var.set(f"错误: {str(e)}")

    def generate_plc(self):
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment

        template_file = self.template_path.get()
        output_file = self.output_path.get()
        if not os.path.exists(template_file):
            messagebox.showerror("错误", "模板文件不存在!")
            return
        if not output_file.endswith('.xlsx'):
            output_file += '.xlsx'

        alias_file = self.alias_file.get() if self.alias_file.get() else None
        if alias_file and not os.path.exists(alias_file):
            alias_file = None
        alias_map = load_alias_map(alias_file)

        try:
            self.status_var.set("正在生成...")
            device_definitions = parse_template(template_file)
            i_points, q_points, iw_points, qw_points = expand_points(device_definitions)

            wb = Workbook()
            ws = wb.active
            ws.title = "PLC点位"
            headers = ['I地址', 'I变量名', 'I别名', 'Q地址', 'Q变量名', 'Q别名', 'IW地址', 'IW变量名', 'IW别名', 'QW地址', 'QW变量名', 'QW别名']
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

            i_start = self.i_start.get()
            q_start = self.q_start.get()
            iw_start = self.iw_start.get()
            qw_start = self.qw_start.get()

            for idx, point in enumerate(i_points):
                ws.cell(row=idx+2, column=1, value=octal_bit('I', i_start, idx))
                ws.cell(row=idx+2, column=2, value=generate_var_name(point))
                ws.cell(row=idx+2, column=3, value=generate_alias(point, alias_map))
            for idx, point in enumerate(q_points):
                ws.cell(row=idx+2, column=4, value=octal_bit('Q', q_start, idx))
                ws.cell(row=idx+2, column=5, value=generate_var_name(point))
                ws.cell(row=idx+2, column=6, value=generate_alias(point, alias_map))
            for idx, point in enumerate(iw_points):
                ws.cell(row=idx+2, column=7, value=octal_word('IW', iw_start, idx, 2))
                ws.cell(row=idx+2, column=8, value=generate_var_name(point))
                ws.cell(row=idx+2, column=9, value=generate_alias(point, alias_map))
            for idx, point in enumerate(qw_points):
                ws.cell(row=idx+2, column=10, value=octal_word('QW', qw_start, idx, 2))
                ws.cell(row=idx+2, column=11, value=generate_var_name(point))
                ws.cell(row=idx+2, column=12, value=generate_alias(point, alias_map))

            max_rows = max(len(i_points), len(q_points), len(iw_points), len(qw_points)) + 2
            for col_idx in range(1, 13):
                max_len = max(len(str(ws.cell(row=r, column=col_idx).value or ''))
                              for r in range(1, max_rows))
                col_letter = chr(64 + col_idx) if col_idx <= 26 else 'A' + chr(64 + col_idx - 26)
                ws.column_dimensions[col_letter].width = max_len + 3

            wb.save(output_file)
            alias_info = f"\n别名映射: 自定义" if alias_file else "\n别名映射: 默认"
            self.status_var.set("生成完成!")
            messagebox.showinfo("成功", f"PLC点位表已生成!{alias_info}\n\n文件: {output_file}\n\nI: {len(i_points)} 个\nQ: {len(q_points)} 个\nIW: {len(iw_points)} 个\nQW: {len(qw_points)} 个")
        except Exception as e:
            self.status_var.set(f"错误: {str(e)}")
            messagebox.showerror("错误", f"生成失败:\n{str(e)}")

    # ══════════════════════════════════════════════════════
    # 页面2: 模块排序生成
    # ══════════════════════════════════════════════════════
    def _build_module_page(self):
        self.status_var.set("就绪")
        f = ttk.Frame(self.page_container, padding=20)
        f.pack(fill=tk.BOTH, expand=True)
        self.pages['module'] = f

        # 参数配置区域 (紧凑横向布局)
        param_frame = ttk.LabelFrame(f, text="模块参数设置", padding=12, style='Card.TLabelframe')
        param_frame.pack(fill=tk.X, pady=(0, 10))


        def make_module_row(parent, title, count_var, start_var, start_label, point_total, col=0):
            sub = ttk.LabelFrame(parent, text=title, padding=8)
            sub.grid(row=0, column=col, padx=4, pady=2, sticky=tk.W)
            ttk.Label(sub, text="数量:").grid(row=0, column=0, sticky=tk.W)
            ttk.Spinbox(sub, from_=1, to=32, textvariable=count_var, width=6).grid(row=0, column=1, padx=4)
            ttk.Label(sub, text=start_label).grid(row=0, column=2, sticky=tk.W, padx=(8, 0))
            ttk.Spinbox(sub, from_=0, to=1000, textvariable=start_var, width=6).grid(row=0, column=3, padx=4)

        # 模块变量
        self.mod_i_count = tk.IntVar(value=4)
        self.mod_i_start = tk.IntVar(value=2)
        self.mod_q_count = tk.IntVar(value=4)
        self.mod_q_start = tk.IntVar(value=2)
        self.mod_iw_count = tk.IntVar(value=2)
        self.mod_iw_start = tk.IntVar(value=10)
        self.mod_iw_eight_count = tk.IntVar(value=2)  # 前N个IW为8点，默认全部8点
        self.mod_qw_count = tk.IntVar(value=2)
        self.mod_qw_start = tk.IntVar(value=16)

        grid = ttk.Frame(param_frame)
        grid.pack(fill=tk.X)

        make_module_row(grid, "I 模块", self.mod_i_count, self.mod_i_start, "起始字节:", "", 0)
        make_module_row(grid, "Q 模块", self.mod_q_count, self.mod_q_start, "起始字节:", "", 1)
        make_module_row(grid, "IW 模块", self.mod_iw_count, self.mod_iw_start, "起始偏移:", "", 2)
        make_module_row(grid, "QW 模块", self.mod_qw_count, self.mod_qw_start, "起始偏移:", "", 3)

        # IW 8点/4点 配置行
        iw_config_row = ttk.Frame(param_frame)
        iw_config_row.pack(fill=tk.X, pady=(6, 0))
        ttk.Label(iw_config_row, text="IW前N个为8点:", font=('Microsoft YaHei', 9)).pack(side=tk.LEFT)
        ttk.Spinbox(iw_config_row, from_=0, to=32, textvariable=self.mod_iw_eight_count,
                     width=5).pack(side=tk.LEFT, padx=4)
        self.iw_config_label = tk.Label(iw_config_row, text="(剩余为4点)", font=('Microsoft YaHei', 9),
                                         fg=self.c['p2'], bg=self.c['bg'])
        self.iw_config_label.pack(side=tk.LEFT, padx=4)

        # PLC类型 + 输出 + 操作 (一行)
        row2 = ttk.Frame(f)
        row2.pack(fill=tk.X, pady=(0, 10))

        ttk.Label(row2, text="PLC类型:").pack(side=tk.LEFT)
        self.mod_plc_type = tk.StringVar(value="200SMART")
        ttk.Radiobutton(row2, text="S7-1500", variable=self.mod_plc_type, value="1500",
                        command=self.module_preview).pack(side=tk.LEFT, padx=4)
        ttk.Radiobutton(row2, text="200SMART", variable=self.mod_plc_type, value="200SMART",
                        command=self.module_preview).pack(side=tk.LEFT, padx=4)

        ttk.Label(row2, text="  输出文件:").pack(side=tk.LEFT, padx=(20, 0))
        self.module_output = tk.StringVar(value="PLC模块排序表.xlsx")
        ttk.Entry(row2, textvariable=self.module_output, width=25).pack(side=tk.LEFT, padx=6)
        ttk.Button(row2, text="浏览", command=self.browse_module_output).pack(side=tk.LEFT)

        ttk.Button(row2, text="预览", command=self.module_preview).pack(side=tk.LEFT, padx=(20, 4))
        ttk.Button(row2, text="生成排序表", command=self.generate_module, width=14).pack(side=tk.LEFT)

        # 预览
        pf = ttk.LabelFrame(f, text="预览", padding=8, style='Card.TLabelframe')
        pf.pack(fill=tk.BOTH, expand=True)


        self.module_preview_text = tk.Text(pf, height=10, font=("Consolas", 10),
                                           bg=self.c['card'], fg=self.c['p1'],
                                           relief='flat', borderwidth=0)
        self.module_preview_text.pack(fill=tk.BOTH, expand=True)
        self.module_preview()

    def browse_module_output(self):
        fn = filedialog.asksaveasfilename(title="保存位置", defaultextension=".xlsx",
                                          filetypes=[("Excel", "*.xlsx")])
        if fn:
            self.module_output.set(fn)

    def module_preview(self):
        plc_type = self.mod_plc_type.get()
        i_pts = generate_i_modules(self.mod_i_count.get(), self.mod_i_start.get())
        q_pts = generate_q_modules(self.mod_q_count.get(), self.mod_q_start.get())
        iw_eight = self.mod_iw_eight_count.get()
        iw_pts, iw_pts_per_module = generate_iw_modules(
            self.mod_iw_count.get(), self.mod_iw_start.get(), plc_type, iw_eight)
        qw_pts = generate_qw_modules(self.mod_qw_count.get(), self.mod_qw_start.get(), plc_type)

        iw_8_count = min(self.mod_iw_count.get(), iw_eight)
        iw_4_count = self.mod_iw_count.get() - iw_8_count

        preview = (f"【PLC类型】{plc_type}\n\n"
                   f"【I模块】{len(i_pts)}点 ({self.mod_i_count.get()}模块，起始I{self.mod_i_start.get()}.0)\n"
                   f"  {', '.join(i_pts[:8])}...\n\n"
                   f"【Q模块】{len(q_pts)}点 ({self.mod_q_count.get()}模块，起始Q{self.mod_q_start.get()}.0)\n"
                   f"  {', '.join(q_pts[:8])}...\n\n"
                   f"【IW模块】{len(iw_pts)}点 ({self.mod_iw_count.get()}模块，{iw_8_count}个8点+{iw_4_count}个4点，起始IW{self.mod_iw_start.get()})\n"
                   f"  {', '.join(iw_pts[:16])}{'...' if len(iw_pts) > 16 else ''}\n\n"
                   f"【QW模块】{len(qw_pts)}点 ({self.mod_qw_count.get()}模块，起始QW{self.mod_qw_start.get()})\n"
                   f"  {', '.join(qw_pts)}")
        self.module_preview_text.delete(1.0, tk.END)
        self.module_preview_text.insert(1.0, preview)

    def generate_module(self):
        try:
            output = self.module_output.get()
            if not output.endswith('.xlsx'):
                output += '.xlsx'
            generate_module_excel(
                self.mod_i_count.get(), self.mod_i_start.get(),
                self.mod_q_count.get(), self.mod_q_start.get(),
                self.mod_iw_count.get(), self.mod_iw_start.get(),
                self.mod_qw_count.get(), self.mod_qw_start.get(),
                output, plc_type=self.mod_plc_type.get(),
                iw_eight_count=self.mod_iw_eight_count.get()
            )
            self.status_var.set("生成完成!")
            messagebox.showinfo("成功", f"已生成: {output}")
        except Exception as e:
            self.status_var.set(f"错误: {str(e)}")
            messagebox.showerror("错误", str(e))

    # ══════════════════════════════════════════════════════
    # 页面3: 端子表生成
    # ══════════════════════════════════════════════════════
    def _build_terminal_page(self):
        self.status_var.set("就绪")
        f = ttk.Frame(self.page_container, padding=20)
        f.pack(fill=tk.BOTH, expand=True)
        self.pages['terminal'] = f

        gf1 = ttk.LabelFrame(f, text="文件设置", padding=12, style='Card.TLabelframe')
        gf1.pack(fill=tk.X, pady=(0, 10))

        r0 = ttk.Frame(gf1)
        r0.pack(fill=tk.X, pady=3)
        ttk.Label(r0, text="PLC点位表:").pack(side=tk.LEFT)
        self.terminal_input = tk.StringVar(value="PLC点位表.xlsx")
        ttk.Entry(r0, textvariable=self.terminal_input, width=45).pack(side=tk.LEFT, padx=8)
        ttk.Button(r0, text="浏览", command=self.browse_terminal_input).pack(side=tk.LEFT)

        r1 = ttk.Frame(gf1)
        r1.pack(fill=tk.X, pady=3)
        ttk.Label(r1, text="输出文件:").pack(side=tk.LEFT)
        self.terminal_output = tk.StringVar(value="端子表_生成.xlsx")
        ttk.Entry(r1, textvariable=self.terminal_output, width=45).pack(side=tk.LEFT, padx=8)
        ttk.Button(r1, text="浏览", command=self.browse_terminal_output).pack(side=tk.LEFT)

        gf2 = ttk.LabelFrame(f, text="PLC类型与电源设置", padding=12, style='Card.TLabelframe')
        gf2.pack(fill=tk.X, pady=(0, 10))

        r2 = ttk.Frame(gf2)
        r2.pack(fill=tk.X)
        ttk.Label(r2, text="PLC类型:").pack(side=tk.LEFT)
        self.plc_type = tk.StringVar(value="200SMART")
        ttk.Radiobutton(r2, text="S7-1500", variable=self.plc_type, value="1500").pack(side=tk.LEFT, padx=4)
        ttk.Radiobutton(r2, text="200SMART", variable=self.plc_type, value="200SMART").pack(side=tk.LEFT, padx=4)

        ttk.Label(r2, text="  24V+标识:").pack(side=tk.LEFT, padx=(20, 0))
        self.power_pos = tk.StringVar(value="24V+")
        ttk.Entry(r2, textvariable=self.power_pos, width=10).pack(side=tk.LEFT, padx=4)
        ttk.Label(r2, text="24V-标识:").pack(side=tk.LEFT, padx=(8, 0))
        self.power_neg = tk.StringVar(value="24V-")
        ttk.Entry(r2, textvariable=self.power_neg, width=10).pack(side=tk.LEFT, padx=4)

        # IW/QW 前缀 + KA起始
        r_prefix = ttk.Frame(gf2)
        r_prefix.pack(fill=tk.X, pady=(6, 0))
        ttk.Label(r_prefix, text="IW前缀:").pack(side=tk.LEFT)
        self.iw_prefix = tk.StringVar(value="")
        ttk.Entry(r_prefix, textvariable=self.iw_prefix, width=10).pack(side=tk.LEFT, padx=4)
        ttk.Label(r_prefix, text="QW前缀:", foreground=self.c['p2']).pack(side=tk.LEFT, padx=(16, 0))
        self.qw_prefix = tk.StringVar(value="")
        ttk.Entry(r_prefix, textvariable=self.qw_prefix, width=10).pack(side=tk.LEFT, padx=4)
        ttk.Label(r_prefix, text="KA起始:", foreground=self.c['p2']).pack(side=tk.LEFT, padx=(16, 0))
        self.ka_start = tk.StringVar(value="1")
        ttk.Entry(r_prefix, textvariable=self.ka_start, width=6).pack(side=tk.LEFT, padx=4)
        ttk.Label(r_prefix, text="KA加24V:", foreground=self.c['p2']).pack(side=tk.LEFT, padx=(16, 0))
        self.ka_power_ranges = tk.StringVar(value="")
        ttk.Entry(r_prefix, textvariable=self.ka_power_ranges, width=12).pack(side=tk.LEFT, padx=4)
        ttk.Label(r_prefix, text="(如 1~3,5~7)", foreground=self.c['p2']).pack(side=tk.LEFT, padx=4)

        r_iw_power = ttk.Frame(gf2)
        r_iw_power.pack(fill=tk.X, pady=(6, 0))
        ttk.Label(r_iw_power, text="IW串电范围:").pack(side=tk.LEFT)
        self.iw_power_ranges = tk.StringVar(value="")
        ttk.Entry(r_iw_power, textvariable=self.iw_power_ranges, width=12).pack(side=tk.LEFT, padx=4)
        ttk.Label(r_iw_power, text="(如 1~3,5~7，IW通道序号)", foreground=self.c['p2']).pack(side=tk.LEFT, padx=4)

        # 操作按钮
        btn_row = ttk.Frame(f)
        btn_row.pack(fill=tk.X, pady=(0, 10))
        ttk.Button(btn_row, text="刷新预览", command=self.terminal_preview).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(btn_row, text="生成端子表", command=self.generate_terminal, width=15).pack(side=tk.LEFT)

        gf3 = ttk.LabelFrame(f, text="设备分组预览", padding=8, style='Card.TLabelframe')
        gf3.pack(fill=tk.BOTH, expand=True)


        cols = ("设备名", "I点数", "Q点数", "IW点数", "QW点数", "Q地址")
        self.terminal_tree = ttk.Treeview(gf3, columns=cols, show="headings", height=9)
        widths = [80, 60, 60, 60, 60, 120]
        for col, w in zip(cols, widths):
            self.terminal_tree.heading(col, text=col)
            self.terminal_tree.column(col, width=w, anchor=tk.CENTER)
        scroll = ttk.Scrollbar(gf3, orient=tk.VERTICAL, command=self.terminal_tree.yview)
        self.terminal_tree.configure(yscrollcommand=scroll.set)
        self.terminal_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scroll.pack(side=tk.RIGHT, fill=tk.Y)

        self.terminal_preview()

    def browse_terminal_input(self):
        fn = filedialog.askopenfilename(title="选择PLC点位表", filetypes=[("Excel", "*.xlsx")])
        if fn:
            self.terminal_input.set(fn)
            self.terminal_preview()

    def browse_terminal_output(self):
        fn = filedialog.asksaveasfilename(title="保存位置", defaultextension=".xlsx",
                                          filetypes=[("Excel", "*.xlsx")])
        if fn:
            self.terminal_output.set(fn)

    def terminal_preview(self):
        for item in self.terminal_tree.get_children():
            self.terminal_tree.delete(item)
        path = self.terminal_input.get()
        if not os.path.exists(path):
            self.status_var.set(f"文件不存在: {path}")
            return
        try:
            i_pts, q_pts, iw_pts, qw_pts = parse_plc_points(path)
            devices = group_devices_by_type(i_pts, q_pts, iw_pts, qw_pts)
            for device in sorted(devices.keys()):
                pts = devices[device]
                q_addrs = list(pts['Q'].values()) if pts['Q'] else ['-']
                self.terminal_tree.insert("", tk.END, values=(
                    device, len(pts['I']), len(pts['Q']),
                    len(pts['IW']), len(pts['QW']), ", ".join(q_addrs[:3])
                ))
            self.status_var.set(f"就绪 - 共 {len(devices)} 个设备")
        except Exception as e:
            self.status_var.set(f"错误: {str(e)}")

    def generate_terminal(self):
        input_file = self.terminal_input.get()
        output_file = self.terminal_output.get()
        if not os.path.exists(input_file):
            messagebox.showerror("错误", "PLC点位表不存在!")
            return
        if not output_file.endswith('.xlsx'):
            output_file += '.xlsx'

        plc_type = self.plc_type.get()
        power_pos = self.power_pos.get() or "24V+"
        power_neg = self.power_neg.get() or "24V-"
        iw_prefix = self.iw_prefix.get() or ""
        qw_prefix = self.qw_prefix.get() or ""
        try:
            ka_start = int(self.ka_start.get())
        except:
            ka_start = 1
        ka_power_ranges = self.ka_power_ranges.get() or ""
        iw_power_ranges = self.iw_power_ranges.get() or ""

        try:
            self.status_var.set("正在生成...")
            generate_terminal_excel(input_file, output_file, plc_type=plc_type,
                                    power_pos=power_pos, power_neg=power_neg,
                                    iw_prefix=iw_prefix, qw_prefix=qw_prefix,
                                    ka_start=ka_start, ka_power_ranges=ka_power_ranges,
                                    iw_power_ranges=iw_power_ranges)
            self.status_var.set("生成完成!")
            has_custom = load_priority_json() is not None
            priority_info = "\n优先级: 自定义" if has_custom else "\n优先级: 默认"
            messagebox.showinfo("成功", f"端子表已生成!\nPLC类型: {plc_type}{priority_info}\n\n文件: {output_file}")
        except Exception as e:
            self.status_var.set(f"错误: {str(e)}")
            messagebox.showerror("错误", str(e))

    # ══════════════════════════════════════════════════════
    # 页面4: EPLAN点位表生成
    # ══════════════════════════════════════════════════════
    def _build_eplan_page(self):
        self.status_var.set("就绪")
        f = ttk.Frame(self.page_container, padding=20)
        f.pack(fill=tk.BOTH, expand=True)
        self.pages['eplan'] = f

        # 参数区域 (2x2 布局)
        param_frame = ttk.LabelFrame(f, text="模块参数设置", padding=12, style='Card.TLabelframe')
        param_frame.pack(fill=tk.X, pady=(0, 10))


        self.eplan_i_module = tk.IntVar(value=4)
        self.eplan_i_start = tk.IntVar(value=2)
        self.eplan_q_module = tk.IntVar(value=4)
        self.eplan_q_start = tk.IntVar(value=2)
        self.eplan_iw_module = tk.IntVar(value=2)
        self.eplan_iw_start = tk.IntVar(value=10)
        self.eplan_iw_eight = tk.IntVar(value=2)  # 前N个IW为8点
        self.eplan_qw_module = tk.IntVar(value=2)
        self.eplan_qw_start = tk.IntVar(value=16)

        def make_eplan_row(parent, title, count_var, start_var, start_label):
            row = ttk.Frame(parent)
            row.pack(fill=tk.X, pady=2)
            ttk.Label(row, text=title, width=14, anchor=tk.W).pack(side=tk.LEFT)
            ttk.Label(row, text="模块数量:").pack(side=tk.LEFT, padx=(8, 0))
            ttk.Spinbox(row, from_=1, to=32, textvariable=count_var, width=6).pack(side=tk.LEFT, padx=4)
            ttk.Label(row, text=start_label).pack(side=tk.LEFT, padx=(12, 0))
            ttk.Spinbox(row, from_=0, to=1000, textvariable=start_var, width=6).pack(side=tk.LEFT, padx=4)
            pts = tk.StringVar()
            ttk.Label(row, textvariable=pts, foreground=self.c['p2']).pack(side=tk.LEFT, padx=8)

            def update_pts(*_):
                pts.set(f"(共 {count_var.get() * 16} 点)")
            count_var.trace_add('write', update_pts)
            update_pts()
            return row

        make_eplan_row(param_frame, "I 模块", self.eplan_i_module, self.eplan_i_start, "起始字节:")
        make_eplan_row(param_frame, "Q 模块", self.eplan_q_module, self.eplan_q_start, "起始字节:")
        make_eplan_row(param_frame, "IW 模块", self.eplan_iw_module, self.eplan_iw_start, "起始偏移:")
        # IW 8点/4点 配置
        iw_cfg = ttk.Frame(param_frame)
        iw_cfg.pack(fill=tk.X, pady=2)
        ttk.Label(iw_cfg, text="IW前N个为8点:", width=14, anchor=tk.W).pack(side=tk.LEFT)
        ttk.Spinbox(iw_cfg, from_=0, to=32, textvariable=self.eplan_iw_eight, width=6).pack(side=tk.LEFT, padx=4)
        ttk.Label(iw_cfg, text="(剩余为4点)", foreground=self.c['p2']).pack(side=tk.LEFT, padx=4)

        make_eplan_row(param_frame, "QW 模块", self.eplan_qw_module, self.eplan_qw_start, "起始偏移:")

        # 输出设置
        out_frame = ttk.LabelFrame(f, text="输出设置", padding=12, style='Card.TLabelframe')
        out_frame.pack(fill=tk.X, pady=(0, 10))


        r_out = ttk.Frame(out_frame)
        r_out.pack(fill=tk.X)
        ttk.Label(r_out, text="输出文件:").pack(side=tk.LEFT)
        self.eplan_output = tk.StringVar(value="EPLAN点位表.xlsx")
        ttk.Entry(r_out, textvariable=self.eplan_output, width=45).pack(side=tk.LEFT, padx=8)
        ttk.Button(r_out, text="浏览", command=self.browse_eplan_output).pack(side=tk.LEFT)

        # 按钮
        btn_row = ttk.Frame(f)
        btn_row.pack(fill=tk.X)
        ttk.Button(btn_row, text="生成EPLAN点位表", command=self.generate_eplan,
                   width=20).pack(side=tk.LEFT)

        # 信息提示
        tip = ttk.LabelFrame(f, text="说明", padding=12, style='Card.TLabelframe')
        tip.pack(fill=tk.BOTH, expand=True, pady=(10, 0))

        info = ("I/Q 模块: 每组 16 点，偶数位在前(.0,.2,.4,.6)，奇数位在后(.1,.3,.5,.7)\n"
                "IW 模块: 每组 8 点，交错间隔 4 (起始,+4,+8,+12,+2,+6,+10,+14)\n"
                "QW 模块: 每组 4 点，交错排列 (起始,+4,+2,+6)\n"
                "Q 模块附带 KA 继电器编号")
        ttk.Label(tip, text=info, font=('Microsoft YaHei', 9),
                  foreground=self.c['p2'], justify=tk.LEFT).pack(anchor=tk.W, padx=4, pady=4)

    def browse_eplan_output(self):
        fn = filedialog.asksaveasfilename(title="保存位置", defaultextension=".xlsx",
                                          filetypes=[("Excel", "*.xlsx")])
        if fn:
            self.eplan_output.set(fn)

    def generate_eplan(self):
        try:
            output_file = self.eplan_output.get()
            if not output_file.endswith('.xlsx'):
                output_file += '.xlsx'
            self.status_var.set("正在生成...")
            generate_eplan_sheet(
                module_i=self.eplan_i_module.get(), start_i=self.eplan_i_start.get(),
                module_q=self.eplan_q_module.get(), start_q=self.eplan_q_start.get(),
                module_iw=self.eplan_iw_module.get(), start_iw=self.eplan_iw_start.get(),
                module_qw=self.eplan_qw_module.get(), start_qw=self.eplan_qw_start.get(),
                output_file=output_file, iw_eight_count=self.eplan_iw_eight.get()
            )
            self.status_var.set("生成完成!")
            messagebox.showinfo("成功", f"EPLAN点位表已生成!\n\n文件: {output_file}")
        except Exception as e:
            self.status_var.set(f"错误: {str(e)}")
            messagebox.showerror("错误", f"生成失败:\n{str(e)}")


if __name__ == '__main__':
    root = tk.Tk()
    app = PLCGeneratorApp(root)
    root.mainloop()
