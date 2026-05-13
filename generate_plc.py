# -*- coding: utf-8 -*-
"""PLC点位表自动生成工具"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import re


# 默认别名映射表 {(设备类型, 地址类型): 别名模板}
DEFAULT_ALIAS_MAP = {
    # I点位映射 (输入)
    ('电机', 'I'): "D_电机.{device}.{point} := {var};",
    ('单线圈', 'I'): "D_单线圈.{device}.{point} := {var};",
    ('双线圈', 'I'): "D_双线圈.{device}.{point} := {var};",
    ('电磁阀', 'I'): "D_电磁阀.{device}.{point} := {var};",
    ('阀门', 'I'): "D_阀门.{device}.{point} := {var};",
    # Q点位映射 (输出)
    ('电机', 'Q'): "{var} := D_电机.{device}.{point};",
    ('单线圈', 'Q'): "{var} := D_单线圈.{device}.{point};",
    ('双线圈', 'Q'): "{var} := D_双线圈.{device}.{point};",
    ('电磁阀', 'Q'): "{var} := D_电磁阀.{device}.{point};",
    ('阀门', 'Q'): "{var} := D_阀门.{device}.{point};",
    # IW/QW映射
    ('*', 'IW'): "D_模拟量读取.{var}.AIW := {var};",
    ('*', 'QW'): "{var} := D_模拟量写入.{var}.AQW;",
}


def load_alias_map(file_path=None):
    """从Excel加载别名映射表，返回 {(设备类型, 地址类型): 别名模板}"""
    if file_path is None:
        return dict(DEFAULT_ALIAS_MAP)
    
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path)
        ws = wb.active
        
        alias_map = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            dtype, addr_type, template = row[0], row[1], row[2]
            if dtype and addr_type and template:
                alias_map[(str(dtype), str(addr_type))] = str(template)
        
        if alias_map:
            return alias_map
    except Exception as e:
        print(f"加载别名映射表失败: {e}")
    
    return dict(DEFAULT_ALIAS_MAP)


def export_alias_map(file_path):
    """导出默认别名映射表到Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    
    wb = Workbook()
    ws = wb.active
    ws.title = "别名映射表"
    
    headers = ['设备类型', '地址类型(I/Q/IW/QW)', '别名模板(用{device}{point}{var}占位)']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # 写入默认数据
    for idx, ((dtype, addr_type), template) in enumerate(DEFAULT_ALIAS_MAP.items(), 2):
        ws.cell(row=idx, column=1, value=dtype)
        ws.cell(row=idx, column=2, value=addr_type)
        ws.cell(row=idx, column=3, value=template)
    
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 45
    
    wb.save(file_path)
    print(f"别名映射表已导出: {file_path}")


def parse_template(file_path):
    """解析模板，提取所有设备点位"""
    from openpyxl import load_workbook
    wb = load_workbook(file_path)
    ws = wb['Sheet1']
    
    rows = list(ws.iter_rows(values_only=True))
    
    device_types = {}
    current_type = None
    
    for row in rows:
        name, col_i, col_q, col_iw, col_qw = row
        
        # 检测设备类型行（以"设备_"开头）
        if name and name.startswith('设备_'):
            current_type = name[3:].strip()
            continue
        
        if name is not None:
            device_name = name.strip()
            
            if current_type not in device_types:
                device_types[current_type] = {
                    'devices': [],
                    'i_templates': [],  # 不排序，保持模板顺序
                    'q_templates': [],
                    'iw_templates': [],
                    'qw_templates': []
                }
            
            device_types[current_type]['devices'].append(device_name)
            
            # 保持模板中的顺序
            if col_i and col_i not in device_types[current_type]['i_templates']:
                device_types[current_type]['i_templates'].append(col_i)
            if col_q and col_q not in device_types[current_type]['q_templates']:
                device_types[current_type]['q_templates'].append(col_q)
            if col_iw and col_iw not in device_types[current_type]['iw_templates']:
                device_types[current_type]['iw_templates'].append(col_iw)
            if col_qw and col_qw not in device_types[current_type]['qw_templates']:
                device_types[current_type]['qw_templates'].append(col_qw)
        else:
            # 独立点位行，保持顺序
            if current_type and current_type in device_types:
                if col_i and col_i not in device_types[current_type]['i_templates']:
                    device_types[current_type]['i_templates'].append(col_i)
                if col_q and col_q not in device_types[current_type]['q_templates']:
                    device_types[current_type]['q_templates'].append(col_q)
    
    return device_types


def expand_points(device_definitions):
    """展开点位"""
    i_points, q_points, iw_points, qw_points = [], [], [], []
    
    for dtype in device_definitions:
        defs = device_definitions[dtype]
        
        for device in defs['devices']:
            # 按模板顺序，添加设备类型和地址类型
            for point in defs['i_templates']:
                i_points.append({'device': device, 'point': point, 'dtype': dtype, 'addr_type': 'I'})
            for point in defs['q_templates']:
                q_points.append({'device': device, 'point': point, 'dtype': dtype, 'addr_type': 'Q'})
            for point in defs['iw_templates']:
                iw_points.append({'device': device, 'point': point, 'dtype': dtype, 'addr_type': 'IW'})
            for point in defs['qw_templates']:
                qw_points.append({'device': device, 'point': point, 'dtype': dtype, 'addr_type': 'QW'})
    
    return i_points, q_points, iw_points, qw_points


def bit_address(prefix, start_byte, index):
    """生成位地址（十进制）"""
    byte = start_byte + index // 8
    bit = index % 8
    return f"{prefix}{byte}.{bit}"


def word_address(prefix, start, index, interval=2):
    """生成字地址，间隔可设置"""
    return f"{prefix}{start + index * interval}"

# 别名（兼容旧名称）
octal_bit = bit_address
octal_word = word_address

def generate_var_name(point):
    return f"{point['device']}{point['point']}"


def generate_alias(point, alias_map=None):
    """生成别名/赋值语句
    alias_map: 别名映射表 {(设备类型, 地址类型): 模板字符串}
    """
    if alias_map is None:
        alias_map = DEFAULT_ALIAS_MAP
    
    device = point['device']
    point_name = point['point']
    var_name = f"{device}{point_name}"
    dtype = point['dtype']
    addr_type = point['addr_type']
    
    # 查找匹配的模板：优先精确匹配，其次通配符*
    template = alias_map.get((dtype, addr_type))
    if template is None:
        template = alias_map.get(('*', addr_type))
    
    if template:
        return template.format(device=device, point=point_name, var=var_name)
    
    # 没有匹配模板时返回变量名
    return f"{var_name};"


def generate_plc_sheet(template_file, output_file, i_start=0, q_start=0, iw_start=0, qw_start=0, alias_file=None):
    """生成PLC点位表到新文件
    alias_file: 别名映射表Excel文件路径
    """
    device_definitions = parse_template(template_file)
    alias_map = load_alias_map(alias_file)
    
    print("设备类型定义:")
    for dtype, defs in device_definitions.items():
        print(f"  {dtype}: {len(defs['devices'])}个 {defs['devices']}")
        print(f"    I={defs['i_templates']}, Q={defs['q_templates']}")
        print(f"    IW={defs['iw_templates']}, QW={defs['qw_templates']}")
    
    i_points, q_points, iw_points, qw_points = expand_points(device_definitions)
    
    print(f"\n展开后的点位: I={len(i_points)}, Q={len(q_points)}, IW={len(iw_points)}, QW={len(qw_points)}")
    print(f"别名映射表: {'已加载自定义' if alias_file else '使用默认'}")
    
    # 创建新工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "PLC点位"
    
    # 表头
    headers = ['I地址', 'I变量名', 'I别名', 'Q地址', 'Q变量名', 'Q别名', 'IW地址', 'IW变量名', 'IW别名', 'QW地址', 'QW变量名', 'QW别名']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # 写入I点位（十进制，I起始.0）
    for idx, point in enumerate(i_points):
        ws.cell(row=idx+2, column=1, value=bit_address('I', i_start, idx))
        ws.cell(row=idx+2, column=2, value=generate_var_name(point))
        ws.cell(row=idx+2, column=3, value=generate_alias(point, alias_map))
    
    # 写入Q点位（十进制，Q起始.0）
    for idx, point in enumerate(q_points):
        ws.cell(row=idx+2, column=4, value=bit_address('Q', q_start, idx))
        ws.cell(row=idx+2, column=5, value=generate_var_name(point))
        ws.cell(row=idx+2, column=6, value=generate_alias(point, alias_map))
    
    # 写入IW点位（间隔2）
    for idx, point in enumerate(iw_points):
        ws.cell(row=idx+2, column=7, value=word_address('IW', iw_start, idx, 2))
        ws.cell(row=idx+2, column=8, value=generate_var_name(point))
        ws.cell(row=idx+2, column=9, value=generate_alias(point, alias_map))
    
    # 写入QW点位（间隔2）
    for idx, point in enumerate(qw_points):
        ws.cell(row=idx+2, column=10, value=word_address('QW', qw_start, idx, 2))
        ws.cell(row=idx+2, column=11, value=generate_var_name(point))
        ws.cell(row=idx+2, column=12, value=generate_alias(point, alias_map))
    
    # 调整列宽
    for col_idx in range(1, 13):
        max_len = max(len(str(ws.cell(row=r, column=col_idx).value or '')) 
                      for r in range(1, max(len(i_points), len(q_points), len(iw_points), len(qw_points)) + 2))
        ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else 'A' + chr(64 + col_idx - 26)].width = max_len + 3
    
    wb.save(output_file)
    print(f"\nPLC点位表已生成: {output_file}")


if __name__ == '__main__':
    template_file = '点表模板.xlsx'
    output_file = 'PLC点位表.xlsx'
    
    generate_plc_sheet(template_file, output_file, i_start=2, q_start=0, iw_start=10, qw_start=0)
